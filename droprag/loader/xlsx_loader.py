"""DropRAG XLSX 加载器 - openpyxl"""

import os
from typing import Optional, List
from droprag.loader import LoaderBase, LoadedDocument, _get_folder_info, _get_file_mtime
from droprag.logging import get_logger

log = get_logger(__name__)


class XlsxLoader(LoaderBase):
    extensions = [".xlsx", ".xls"]

    def load(self, filepath: str, base_path: str) -> Optional[LoadedDocument]:
        try:
            import openpyxl
        except ImportError:
            log.warning("openpyxl 未安装，跳过 XLSX 文件: pip install droprag[office]")
            return None

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            log.debug(f"XLSX 加载失败: {filepath} ({e})")
            return None

        all_sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            header = None

            for row in ws.iter_rows(values_only=True):
                # 过滤空行
                cells = [str(c) if c is not None else "" for c in row]
                if not any(cells):
                    continue

                if header is None:
                    header = cells
                    rows.append(" | ".join(cells))
                else:
                    rows.append(" | ".join(cells))

            if rows:
                sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
                all_sheets.append(sheet_text)

        wb.close()

        if not all_sheets:
            return None

        content = "\n\n".join(all_sheets)
        folder, subfolder = _get_folder_info(filepath, base_path)
        return LoadedDocument(
            content=content,
            source=filepath,
            filename=os.path.basename(filepath),
            file_type="xlsx",
            category="",
            folder=folder,
            subfolder=subfolder,
            file_size=os.path.getsize(filepath),
            file_mtime=_get_file_mtime(filepath),
        )
